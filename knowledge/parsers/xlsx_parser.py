
"""
Excel 文档解析器
"""
from pathlib import Path
from typing import Optional
from .base import BaseParser

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class XlsxParser(BaseParser):
    """Excel (.xlsx, .xls) 文档解析器"""
    
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xls"]
    
    def to_markdown(self, file_path: str, output_path: Optional[str] = None) -> str:
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        markdown_parts = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_parts = [f"## Sheet: {sheet_name}\n"]
            
            # 获取所有行数据
            rows_data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    rows_data.append(row)
            
            if rows_data:
                md_table = self._rows_to_markdown(rows_data)
                sheet_parts.append(md_table)
                markdown_parts.append("\n".join(sheet_parts))
        
        markdown_text = "\n\n".join(markdown_parts)
        
        if output_path:
            self._save_markdown(markdown_text, output_path)
        
        return markdown_text
    
    def _rows_to_markdown(self, rows: list[list]) -> str:
        """将行数据转换为 Markdown 表格"""
        if not rows:
            return ""
        
        md_rows = []
        for row in rows:
            cells = [str(cell).strip().replace('\n', ' ') if cell is not None else '' for cell in row]
            md_rows.append(f"| {' | '.join(cells)} |")
        
        # 添加分隔线
        separator = f"| {' | '.join(['---'] * len(rows[0]))} |"
        md_rows.insert(1, separator)
        
        return "\n".join(md_rows)

